import io
import json
import re
import uuid
from datetime import datetime, timezone

from fastapi import APIRouter, HTTPException, UploadFile, File, Form, BackgroundTasks
from pydantic import BaseModel
from pathlib import Path
from openpyxl import load_workbook
from src.services.curriculum_service import (
    identify_content_gaps,
    process_kb_upload_job,
    process_generate_job,
    regenerate_lesson_content,
    restore_document_version,
)
from src.services.quiz_service import generate_quiz
from src.core.database import DepartmentScopedStore
from src.core.config import DEFAULT_DEPARTMENT, SUPPORTED_MIME_TYPES

router = APIRouter(prefix="/api/kb", tags=["knowledge_base"])

MAX_UPLOAD_BYTES = 20 * 1024 * 1024  # Gemini's inline-request limit (~20MB)

_SPREADSHEET_EXTENSIONS = (".xlsx", ".xls")


def _extract_spreadsheet_text(data: bytes) -> str:
    """Dump every sheet's cell values into a readable text block so a
    spreadsheet upload joins the same chunk/gap-analysis pipeline as any
    other text-family document instead of needing its own content_category."""
    workbook = load_workbook(io.BytesIO(data), data_only=True)
    parts = []
    for sheet in workbook.worksheets:
        parts.append(f"## {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(v) if v is not None else "" for v in row]
            if any(c.strip() for c in cells):
                parts.append(" | ".join(cells))
    return "\n".join(parts)

@router.get("/documents")
async def api_kb_documents(department: str = DEFAULT_DEPARTMENT):
    """List all knowledge base documents."""
    store = DepartmentScopedStore(department)
    docs = store.read_knowledge_base()
    return {"documents": docs, "count": len(docs)}

class ValidateDocumentRequest(BaseModel):
    document_content: str

@router.post("/validate")
async def api_kb_validate(body: ValidateDocumentRequest, department: str = DEFAULT_DEPARTMENT):
    """Validate a document against the knowledge base."""
    result = identify_content_gaps(
        document_content=body.document_content,
        department=department,
    )
    return result

def _require_manager(role: str) -> None:
    if role != "manager":
        raise HTTPException(status_code=403, detail="Only a manager can perform this action.")


class DeleteDocumentRequest(BaseModel):
    role: str = ""
    department: str = DEFAULT_DEPARTMENT

@router.delete("/documents/{filename}")
async def api_delete_kb_document(filename: str, req: DeleteDocumentRequest):
    """Remove an uploaded document entirely (raw file, catalog copy, and its
    parsed chunks) — without this, the same filename can never be re-uploaded,
    since /upload's duplicate check only knows the source document is still on disk."""
    _require_manager(req.role)
    store = DepartmentScopedStore(req.department)

    deleted = store.delete_raw_document(filename)
    store.delete_catalog_input(filename)
    chunks_doc_id = f"{Path(filename).stem}_chunks"
    store.delete_knowledge_document(chunks_doc_id)
    store.delete_version_history(filename)

    if not deleted:
        raise HTTPException(status_code=404, detail=f"Document '{filename}' not found.")

    return {"status": "deleted", "filename": filename}


@router.get("/documents/{filename}/versions")
async def api_kb_document_versions(filename: str, department: str = DEFAULT_DEPARTMENT):
    """List the version history for a document (live + archived entries)."""
    store = DepartmentScopedStore(department)
    entries = store.read_version_history(filename)
    return {"filename": filename, "versions": entries, "count": len(entries)}


class RestoreVersionRequest(BaseModel):
    role: str = ""
    department: str = DEFAULT_DEPARTMENT
    uploaded_by: str = ""


@router.post("/documents/{filename}/versions/{version}/restore")
async def api_kb_restore_version(filename: str, version: int, req: RestoreVersionRequest):
    """Restore an archived version of a document back to current.

    Non-destructive: the content being replaced is itself archived first, so
    restoring never permanently discards whatever was current beforehand.
    """
    _require_manager(req.role)
    try:
        result = restore_document_version(filename, version, req.department, req.uploaded_by)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return result


@router.post("/upload")
async def api_kb_upload(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    department: str = Form(DEFAULT_DEPARTMENT),
    version_action: str = Form(""),
    role: str = Form(""),
    uploaded_by: str = Form(""),
):
    """Accept a document and process it asynchronously.

    The cheap, instant parse checks (filename, extension, UTF-8, empty) run
    synchronously so the client fails fast on obvious errors. If the filename
    already exists and no version_action is chosen, returns a ``duplicate``
    response so the client can prompt Overwrite vs. New Version. Otherwise a
    background job is queued (validate → chunk → save → soft-flag conflicts) and
    a ``job_id`` is returned immediately; the client polls /upload/status/{id}.
    """
    _require_manager(role)
    # ── Parse (synchronous, instant) ─────────────────────────────────────────
    # Narrow file.filename (str | None) immediately so the rest of the function
    # can use the typed `filename: str` variable without repeated None checks.
    if not file.filename:
        raise HTTPException(status_code=400, detail="Uploaded file has no filename.")
    filename: str = file.filename

    ext = Path(filename).suffix.lower()
    if ext not in SUPPORTED_MIME_TYPES:
        supported_list = ", ".join(sorted(SUPPORTED_MIME_TYPES))
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type '{ext}'. Supported types: {supported_list}"
        )
    mime_type, content_category = SUPPORTED_MIME_TYPES[ext]

    content_bytes = await file.read()
    if len(content_bytes) > MAX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=400,
            detail=f"File exceeds the {MAX_UPLOAD_BYTES // (1024 * 1024)}MB upload limit."
        )
    if not content_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    # Text-family documents are decoded to str and keep the original chunking
    # pipeline; everything else (PDF/image/audio/video) stays as raw bytes and
    # is handed to Gemini natively as binary media. Spreadsheets are also
    # "text" category but need extraction rather than a plain UTF-8 decode.
    if content_category == "text" and ext in _SPREADSHEET_EXTENSIONS:
        try:
            content = _extract_spreadsheet_text(content_bytes)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Could not read spreadsheet: {e}")
        if not content.strip():
            raise HTTPException(status_code=400, detail="Uploaded spreadsheet is empty.")
    elif content_category == "text":
        try:
            content = content_bytes.decode("utf-8")
        except UnicodeDecodeError:
            raise HTTPException(status_code=400, detail="File could not be read as UTF-8 text.")
        if not content.strip():
            raise HTTPException(status_code=400, detail="Uploaded file is empty.")
    else:
        content = content_bytes

    # ── Duplicate check ──────────────────────────────────────────────────────
    store = DepartmentScopedStore(department)
    if store.raw_document_exists(filename) and version_action not in ("overwrite", "new_version"):
        return {
            "status": "duplicate",
            "existing_filename": filename,
            "message": (
                f"A file named '{filename}' already exists. "
                "Choose to overwrite it or save as a new version."
            ),
        }

    # ── Queue background processing ──────────────────────────────────────────
    job_id = f"job_{uuid.uuid4().hex[:8]}"
    store.write_kb_job(job_id, {
        "job_id": job_id,
        "status": "pending",
        "stage": "queued",
        "filename": filename,
        "content_category": content_category,
    })
    background_tasks.add_task(
        process_kb_upload_job, job_id, filename, content, department, version_action,
        mime_type, content_category, uploaded_by,
    )
    return {"status": "processing", "job_id": job_id}


@router.get("/upload/status/{job_id}")
async def api_kb_upload_status(job_id: str, department: str = DEFAULT_DEPARTMENT):
    """Return the current status/result of an async KB upload job."""
    store = DepartmentScopedStore(department)
    job = store.read_kb_job(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail=f"Job '{job_id}' not found.")
    return job


@router.get("/conflicts")
async def api_kb_conflicts(status: str | None = "pending", department: str = DEFAULT_DEPARTMENT):
    """List KB conflicts for a department, filtered by status (default pending)."""
    store = DepartmentScopedStore(department)
    conflicts = store.read_conflicts(status=status)
    return {"conflicts": conflicts, "count": len(conflicts)}


class ConflictResolutionRequest(BaseModel):
    resolution: str  # "approved" | "rejected"
    resolved_by: str = "manager"
    notes: str = ""
    department: str = DEFAULT_DEPARTMENT


@router.post("/conflicts/{conflict_id}/resolve")
async def api_kb_resolve_conflict(conflict_id: str, req: ConflictResolutionRequest):
    """Resolve a flagged KB conflict.

    ``approved`` keeps the already-saved document and marks the conflict resolved.
    ``rejected`` dismisses the conflict AND retracts the saved raw document,
    catalog input, and chunks doc that were written during upload.
    """
    store = DepartmentScopedStore(req.department)

    conflict = next(
        (c for c in store.read_conflicts() if c.get("conflict_id") == conflict_id),
        None,
    )
    if conflict is None:
        raise HTTPException(status_code=404, detail=f"Conflict '{conflict_id}' not found.")

    now = datetime.now(timezone.utc).isoformat()

    if req.resolution == "approved":
        conflict["status"] = "resolved"
        conflict["resolved_by"] = req.resolved_by
        conflict["resolved_at"] = now
        conflict["resolution_notes"] = req.notes
        store.write_conflict(conflict_id, conflict)
    elif req.resolution == "rejected":
        conflict["status"] = "dismissed"
        conflict["resolved_by"] = req.resolved_by
        conflict["resolved_at"] = now
        conflict["resolution_notes"] = req.notes
        store.write_conflict(conflict_id, conflict)

        # Retract the saved document that this conflict flagged.
        raw_filename = conflict.get("raw_filename")
        if raw_filename:
            safe_name = raw_filename.replace("/", "_").replace("\\", "_")
            (store.raw_documents_path / safe_name).unlink(missing_ok=True)
            (store.catalog_inputs_path / safe_name).unlink(missing_ok=True)
        chunks_doc_id = conflict.get("chunks_doc_id")
        if chunks_doc_id:
            (store.knowledge_base_path / f"{chunks_doc_id}.json").unlink(missing_ok=True)
    else:
        raise HTTPException(status_code=400, detail="resolution must be 'approved' or 'rejected'")

    return conflict


class GenerateFromInputRequest(BaseModel):
    filename: str = ""
    filenames: list[str] = []
    department: str = DEFAULT_DEPARTMENT
    append_to_latest: bool = False
    manager_id: str = ""
    role: str = ""

@router.post("/generate-from-input")
async def api_generate_from_input(req: GenerateFromInputRequest, background_tasks: BackgroundTasks):
    """Queue background generation of a curriculum (lessons + short quizzes +
    final assessment) from one or more existing files in the catalog/inputs
    directory. Multiple files (req.filenames) are combined into a single course.

    The generated path is saved as the manager's private draft (unofficial path),
    not the department-wide catalog — it only becomes visible to employees once
    the manager explicitly calls POST /learning-path/{path_id}/publish.

    Generation now also pre-builds every lesson's quiz and the course's final
    assessment via Gemini, so this is queued as a background job (like /upload)
    rather than run synchronously — a job_id is returned immediately and the
    client polls GET /generate-status/{job_id}.
    """
    _require_manager(req.role)
    store = DepartmentScopedStore(req.department)

    files = req.filenames if req.filenames else ([req.filename] if req.filename else [])
    if not files:
        raise HTTPException(status_code=400, detail="At least one filename is required.")

    catalog_filenames = {f["filename"] for f in store.list_catalog_inputs()}
    missing = [fn for fn in files if fn not in catalog_filenames]
    if missing:
        raise HTTPException(status_code=404, detail=f"File(s) not found in catalog inputs: {', '.join(missing)}.")

    job_id = f"genjob_{uuid.uuid4().hex[:8]}"
    store.write_kb_job(job_id, {
        "job_id": job_id,
        "status": "pending",
        "stage": "queued",
        "filename": ", ".join(files),
    })
    background_tasks.add_task(
        process_generate_job, job_id, "", req.department, req.append_to_latest, req.manager_id, files
    )
    return {"status": "processing", "job_id": job_id}


@router.get("/generate-status/{job_id}")
async def api_generate_status(job_id: str, department: str = DEFAULT_DEPARTMENT):
    """Return the current status/result of an async course-generation job."""
    store = DepartmentScopedStore(department)
    job = store.read_kb_job(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail=f"Job '{job_id}' not found.")
    return job


class PublishPathRequest(BaseModel):
    manager_id: str
    role: str = ""
    department: str = DEFAULT_DEPARTMENT

@router.post("/learning-path/{path_id}/publish")
async def api_publish_learning_path(path_id: str, req: PublishPathRequest):
    """Promote a manager's draft (unofficial) path to the official, team-wide catalog."""
    _require_manager(req.role)
    store = DepartmentScopedStore(req.department)

    draft = store.read_unofficial_path(req.manager_id, path_id)
    if not draft:
        raise HTTPException(status_code=404, detail=f"Draft '{path_id}' not found for manager '{req.manager_id}'.")

    store.write_standard_path(path_id, draft)
    # Promotion, not duplication — the draft copy must not linger in the
    # unofficial catalog once it's official, or it shows up as a permanent
    # "Awaiting Approval" duplicate of the now-published card.
    store.delete_unofficial_path(req.manager_id, path_id)
    return {"status": "published", "path_id": path_id}


class UpdatePathRequest(BaseModel):
    title: str
    manager_id: str = ""
    role: str = ""
    path_type: str = "official"
    department: str = DEFAULT_DEPARTMENT

@router.patch("/learning-path/{path_id}")
async def api_update_learning_path(path_id: str, req: UpdatePathRequest):
    """Rename a learning path (official or a manager's own draft)."""
    _require_manager(req.role)
    store = DepartmentScopedStore(req.department)

    if req.path_type == "unofficial":
        data = store.read_unofficial_path(req.manager_id, path_id)
        if not data:
            raise HTTPException(status_code=404, detail=f"Draft '{path_id}' not found for manager '{req.manager_id}'.")
        data["title"] = req.title
        store.write_unofficial_path(req.manager_id, path_id, data)
    else:
        data = store.read_standard_path(path_id)
        if not data:
            raise HTTPException(status_code=404, detail=f"Learning path '{path_id}' not found.")
        data["title"] = req.title
        store.write_standard_path(path_id, data)

    # Keep the activated/enrolled copy (used by the path detail page and lesson
    # lookups) in sync so the rename shows up there too, not just in listings.
    activated = store.read_learning_path(path_id)
    if activated:
        activated["title"] = req.title
        store.write_learning_path(path_id, activated)

    return {"status": "updated", "path_id": path_id, "title": req.title}


class DeletePathRequest(BaseModel):
    manager_id: str = ""
    role: str = ""
    path_type: str = "official"
    department: str = DEFAULT_DEPARTMENT

@router.delete("/learning-path/{path_id}")
async def api_delete_learning_path(path_id: str, req: DeletePathRequest):
    """Remove a learning path (official or a manager's own draft) entirely."""
    _require_manager(req.role)
    store = DepartmentScopedStore(req.department)

    if req.path_type == "unofficial":
        deleted = store.delete_unofficial_path(req.manager_id, path_id)
        still_exists_elsewhere = store.read_standard_path(path_id) is not None
    else:
        deleted = store.delete_standard_path(path_id)
        still_exists_elsewhere = store.read_unofficial_path(req.manager_id, path_id) is not None

    if not deleted:
        raise HTTPException(status_code=404, detail=f"Learning path '{path_id}' not found.")

    # Same path_id can briefly exist in both catalogs (e.g. mid-publish) — only
    # drop the shared activated/enrolled copy once neither catalog holds it.
    if not still_exists_elsewhere:
        store.delete_learning_path(path_id)
    return {"status": "deleted", "path_id": path_id}

@router.get("/catalog/inputs")
async def api_catalog_inputs(department: str = DEFAULT_DEPARTMENT):
    """List all input files (uploaded learning materials) in the catalog."""
    store = DepartmentScopedStore(department)
    inputs = store.list_catalog_inputs()
    return {"inputs": inputs, "count": len(inputs)}

@router.get("/catalog/learning-paths")
async def api_catalog_learning_paths(
    department: str = DEFAULT_DEPARTMENT,
    user_id: str | None = None,
    official_only: bool = False,
):
    """List all learning paths (official + unofficial) from the catalog."""
    store = DepartmentScopedStore(department)
    official = store.list_standard_paths()
    unofficial = [] if official_only else store.list_unofficial_paths(user_id=user_id)
    all_paths = official + unofficial
    return {
        "learning_paths": all_paths,
        "count": len(all_paths),
        "official_count": len(official),
        "unofficial_count": len(unofficial),
    }


# ── Manual Path Editing (course titles, lesson content, quiz questions) ──

def _read_catalog_path(store: DepartmentScopedStore, path_id: str, path_type: str, manager_id: str) -> dict | None:
    if path_type == "unofficial":
        return store.read_unofficial_path(manager_id, path_id)
    return store.read_standard_path(path_id)


def _write_catalog_path(store: DepartmentScopedStore, path_id: str, path_type: str, manager_id: str, data: dict) -> None:
    if path_type == "unofficial":
        store.write_unofficial_path(manager_id, path_id, data)
    else:
        store.write_standard_path(path_id, data)


def _find_lesson(data: dict, lesson_id: str) -> dict | None:
    for course in data.get("courses", []):
        for lesson in course.get("lessons", []):
            if lesson.get("lesson_id") == lesson_id:
                return lesson
    return None


@router.get("/learning-path/{path_id}/full")
async def api_get_learning_path_full(
    path_id: str,
    role: str = "",
    manager_id: str = "",
    path_type: str = "official",
    department: str = DEFAULT_DEPARTMENT,
):
    """Return a learning path's full editable structure (courses, lessons,
    and each lesson/course's quiz resolved inline) for the manual-edit page.
    Reads from the catalog copy — the same source api_update_learning_path
    already treats as canonical for edits."""
    _require_manager(role)
    store = DepartmentScopedStore(department)
    data = _read_catalog_path(store, path_id, path_type, manager_id)
    if not data:
        raise HTTPException(status_code=404, detail=f"Learning path '{path_id}' not found.")

    for course in data.get("courses", []):
        for lesson in course.get("lessons", []):
            if lesson.get("short_quiz_id"):
                lesson["quiz"] = store.read_quiz(lesson["short_quiz_id"])
        if course.get("final_assessment_id"):
            course["final_assessment"] = store.read_quiz(course["final_assessment_id"])

    return data


class UpdateCourseTitleRequest(BaseModel):
    title: str
    manager_id: str = ""
    role: str = ""
    path_type: str = "official"
    department: str = DEFAULT_DEPARTMENT


@router.patch("/learning-path/{path_id}/course/{course_id}")
async def api_update_course_title(path_id: str, course_id: str, req: UpdateCourseTitleRequest):
    """Rename a single course within a learning path."""
    _require_manager(req.role)
    store = DepartmentScopedStore(req.department)
    data = _read_catalog_path(store, path_id, req.path_type, req.manager_id)
    if not data:
        raise HTTPException(status_code=404, detail=f"Learning path '{path_id}' not found.")

    course = next((c for c in data.get("courses", []) if c.get("course_id") == course_id), None)
    if not course:
        raise HTTPException(status_code=404, detail=f"Course '{course_id}' not found in path '{path_id}'.")
    course["title"] = req.title
    _write_catalog_path(store, path_id, req.path_type, req.manager_id, data)

    activated = store.read_learning_path(path_id)
    if activated:
        act_course = next((c for c in activated.get("courses", []) if c.get("course_id") == course_id), None)
        if act_course:
            act_course["title"] = req.title
            store.write_learning_path(path_id, activated)

    return {"status": "updated", "path_id": path_id, "course_id": course_id, "title": req.title}


class UpdateLessonRequest(BaseModel):
    title: str
    content: str
    manager_id: str = ""
    role: str = ""
    path_type: str = "official"
    department: str = DEFAULT_DEPARTMENT


@router.patch("/learning-path/{path_id}/lesson/{lesson_id}")
async def api_update_lesson(path_id: str, lesson_id: str, req: UpdateLessonRequest):
    """Update a single lesson's title and body content within a learning path."""
    _require_manager(req.role)
    store = DepartmentScopedStore(req.department)
    data = _read_catalog_path(store, path_id, req.path_type, req.manager_id)
    if not data:
        raise HTTPException(status_code=404, detail=f"Learning path '{path_id}' not found.")

    lesson = _find_lesson(data, lesson_id)
    if not lesson:
        raise HTTPException(status_code=404, detail=f"Lesson '{lesson_id}' not found in path '{path_id}'.")
    lesson["title"] = req.title
    lesson["content"] = req.content
    _write_catalog_path(store, path_id, req.path_type, req.manager_id, data)

    activated = store.read_learning_path(path_id)
    if activated:
        act_lesson = _find_lesson(activated, lesson_id)
        if act_lesson:
            act_lesson["title"] = req.title
            act_lesson["content"] = req.content
            store.write_learning_path(path_id, activated)

    return {"status": "updated", "path_id": path_id, "lesson_id": lesson_id}


class QuizQuestionUpdate(BaseModel):
    question_id: str = ""
    text: str
    options: list[str]
    correct_answer_index: int
    rationale: dict[str, str] = {}
    concept_tags: list[str] = []


class UpdateQuizRequest(BaseModel):
    questions: list[QuizQuestionUpdate]
    role: str = ""
    department: str = DEFAULT_DEPARTMENT


@router.patch("/quiz/{quiz_id}")
async def api_update_quiz(quiz_id: str, req: UpdateQuizRequest):
    """Overwrite a persisted quiz's question set (manager manual edit)."""
    _require_manager(req.role)
    store = DepartmentScopedStore(req.department)
    quiz = store.read_quiz(quiz_id)
    if not quiz:
        raise HTTPException(status_code=404, detail=f"Quiz '{quiz_id}' not found.")

    new_questions = []
    for i, q in enumerate(req.questions):
        qd = q.model_dump()
        if not (0 <= qd["correct_answer_index"] < len(qd["options"])):
            raise HTTPException(status_code=400, detail=f"correct_answer_index out of range for question {i + 1}.")
        if not qd.get("question_id"):
            qd["question_id"] = f"q_{uuid.uuid4().hex[:6]}"
        new_questions.append(qd)

    quiz["questions"] = new_questions
    quiz["question_count"] = len(new_questions)
    store.write_quiz(quiz_id, quiz)
    return {"status": "updated", "quiz_id": quiz_id}


class RegenerateLessonRequest(BaseModel):
    lesson_title: str
    lesson_content: str
    instruction: str = ""
    role: str = ""
    department: str = DEFAULT_DEPARTMENT


@router.post("/lesson/{lesson_id}/regenerate")
async def api_regenerate_lesson(lesson_id: str, req: RegenerateLessonRequest):
    """Ask Gemini to rewrite a lesson's title/content for manager review.
    Returns a draft only — the manager must still hit Save to persist it."""
    _require_manager(req.role)
    return regenerate_lesson_content(
        lesson_title=req.lesson_title,
        lesson_content=req.lesson_content,
        instruction=req.instruction,
    )


class RegenerateQuizRequest(BaseModel):
    topic: str
    difficulty: str = "medium"
    question_count: int = 5
    quiz_type: str = "short_quiz"
    instruction: str = ""
    role: str = ""
    department: str = DEFAULT_DEPARTMENT


@router.post("/quiz/{quiz_id}/regenerate")
async def api_regenerate_quiz(quiz_id: str, req: RegenerateQuizRequest):
    """Ask Gemini to generate a fresh question set for manager review.
    Returns a draft only — the manager must still hit Save (PATCH /quiz/{id}) to persist it."""
    _require_manager(req.role)
    topic = f"{req.topic} — {req.instruction.strip()}" if req.instruction.strip() else req.topic
    draft = generate_quiz(
        topic=topic,
        difficulty=req.difficulty,
        question_count=req.question_count,
        quiz_type=req.quiz_type,
        department=req.department,
    )
    return {"status": "success", "questions": draft["questions"]}
