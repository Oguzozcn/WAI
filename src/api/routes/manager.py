"""Manager Dashboard API Routes.

Provides three KPI bucket endpoints for the manager view:
  - /team-kpis     → Team/Managerial Bucket (aggregated health)
  - /reports       → Individual Bucket (direct reports progress)
  - /strategic     → Strategic Bucket (team vs department baseline)
"""

import io

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.core.database import DepartmentScopedStore, KPIStoreReader
from src.core.config import DEFAULT_DEPARTMENT, AT_RISK_READINESS_THRESHOLD, PASS_THRESHOLD
from src.core.dev_config import get_param
from src.services.reporting_service import ensure_kpi_payload_for_today

router = APIRouter(prefix="/api/manager", tags=["manager"])


def _get_direct_reports(store: DepartmentScopedStore, manager_id: str) -> list[dict]:
    """Return all user progress records where manager_id matches."""
    all_progress = store.read_all_user_progress()
    return [p for p in all_progress if p.get("manager_id") == manager_id]


def _status_label(progress: dict) -> str:
    """Return a human-readable status label for a user's progress."""
    state = progress.get("current_state", "enrolled")
    mapping = {
        "completed": "Completed",
        "passed": "Completed",
        "PENDING_VERIFIED_HUMAN_APPROVAL": "Awaiting Sign-off",
        "enrolled": "Not Started",
        "bypass_locked": "Bypass Locked",
        "failed": "Needs Support",
    }
    if state in mapping:
        return mapping[state]
    # "progressing" states
    return "In Progress"


def _rag_color(readiness_score: float) -> str:
    """Return Red/Amber/Green color based on readiness score."""
    if readiness_score >= PASS_THRESHOLD:
        return "green"
    elif readiness_score >= AT_RISK_READINESS_THRESHOLD:
        return "amber"
    return "red"


def _build_report_rows(reports: list[dict]) -> list[dict]:
    """Shared row-building logic for the /reports JSON endpoint and the CSV
    export, so the two can never drift apart in what a "row" means."""
    max_courses = get_param("MAX_COURSES")
    rows = []
    for p in reports:
        completed_courses = p.get("completed_courses", [])
        completion_rate = round(len(completed_courses) / max_courses * 100, 1)
        readiness = p.get("readiness_score", 0.0)
        rows.append({
            "user_id": p.get("user_id", ""),
            "display_name": p.get("display_name", p.get("user_id", "")),
            "status": _status_label(p),
            "current_course_id": p.get("current_course_id", ""),
            "courses_completed": len(completed_courses),
            "completion_rate_pct": completion_rate,
            "readiness_score": round(readiness, 2),
            "rag_color": _rag_color(readiness),
            "is_at_risk": readiness < AT_RISK_READINESS_THRESHOLD,
        })
    # Sort by readiness ascending — at-risk employees appear at the top.
    rows.sort(key=lambda r: r["readiness_score"])
    return rows


def _formula_safe(value) -> str:
    """Neutralize formula injection (Excel executes cell text starting with
    =, +, -, @ as a formula) by prefixing with a leading apostrophe."""
    text = "" if value is None else str(value)
    if text and text[0] in ("=", "+", "-", "@"):
        return "'" + text
    return text


_RAG_FILLS = {
    "green": PatternFill("solid", fgColor="DCFCE7"),
    "amber": PatternFill("solid", fgColor="FEF3C7"),
    "red": PatternFill("solid", fgColor="FEE2E2"),
}
_RAG_FONTS = {
    "green": Font(color="15803D", bold=True),
    "amber": Font(color="B45309", bold=True),
    "red": Font(color="B91C1C", bold=True),
}


# ── Bucket 1: Team / Managerial KPIs ─────────────────────────────────────────

def _require_manager(role: str) -> None:
    if role != "manager":
        raise HTTPException(status_code=403, detail="Only a manager can view this data.")


@router.get("/{manager_id}/team-kpis")
async def manager_team_kpis(manager_id: str, department: str = DEFAULT_DEPARTMENT, role: str = ""):
    """Team Bucket — aggregated health metrics for all direct reports.

    Returns:
        - team_size, active_learners, completion_rate, avg_readiness_score
        - pass_rate, at_risk_count, top_gap_areas
    """
    _require_manager(role)
    store = DepartmentScopedStore(department)
    reports = _get_direct_reports(store, manager_id)

    if not reports:
        raise HTTPException(
            status_code=404,
            detail=f"No direct reports found for manager '{manager_id}' in department '{department}'.",
        )

    total = len(reports)
    active = sum(
        1 for p in reports
        if p.get("current_state") not in ("enrolled", "completed", "passed", None)
    )
    completed = sum(1 for p in reports if p.get("current_state") in ("completed", "passed"))

    # Completion rate: avg of (completed_courses / MAX_COURSES) per employee
    max_courses = get_param("MAX_COURSES")
    completion_rates = [
        len(p.get("completed_courses", [])) / max_courses for p in reports
    ]
    avg_completion_rate = sum(completion_rates) / total if total else 0.0

    # Readiness
    readiness_scores = [p.get("readiness_score", 0.0) for p in reports]
    avg_readiness = sum(readiness_scores) / total if total else 0.0
    at_risk_count = sum(1 for s in readiness_scores if s < AT_RISK_READINESS_THRESHOLD)

    # Pass rate
    assessed = sum(
        1 for p in reports
        if p.get("current_state") in ("completed", "passed", "failed")
    )
    pass_rate = (completed / assessed * 100) if assessed > 0 else 0.0

    # Top gap areas (de-duplicated concept tags across the team)
    gap_counter: dict[str, int] = {}
    for p in reports:
        for concept, count in p.get("error_retention_matrix", {}).items():
            gap_counter[concept] = gap_counter.get(concept, 0) + count
    top_gaps = sorted(gap_counter, key=gap_counter.get, reverse=True)[:5]  # type: ignore[arg-type]

    return {
        "manager_id": manager_id,
        "department": department,
        "team_size": total,
        "active_learners": active,
        "completed_count": completed,
        "avg_completion_rate_pct": round(avg_completion_rate * 100, 1),
        "avg_readiness_score": round(avg_readiness, 2),
        "at_risk_count": at_risk_count,
        "pass_rate_pct": round(pass_rate, 1),
        "top_gap_areas": top_gaps,
        "rag_status": _rag_color(avg_readiness),
    }


# ── Bucket 2: Individual Progress (Direct Reports) ────────────────────────────

@router.get("/{manager_id}/reports")
async def manager_direct_reports(manager_id: str, department: str = DEFAULT_DEPARTMENT, role: str = ""):
    """Individual Bucket — one row per direct report showing their learning progress.

    Returns per-employee:
        - display_name, status label, current_course_id,
          completion_rate, readiness_score, rag_color
    """
    _require_manager(role)
    store = DepartmentScopedStore(department)
    reports = _get_direct_reports(store, manager_id)

    if not reports:
        raise HTTPException(
            status_code=404,
            detail=f"No direct reports found for manager '{manager_id}' in department '{department}'.",
        )

    rows = _build_report_rows(reports)

    return {
        "manager_id": manager_id,
        "department": department,
        "report_count": len(rows),
        "reports": rows,
    }


@router.get("/{manager_id}/reports/export")
async def manager_export_reports_excel(manager_id: str, department: str = DEFAULT_DEPARTMENT, role: str = ""):
    """Same data as /reports, as a downloadable, formatted Excel workbook."""
    _require_manager(role)
    store = DepartmentScopedStore(department)
    reports = _get_direct_reports(store, manager_id)

    if not reports:
        raise HTTPException(
            status_code=404,
            detail=f"No direct reports found for manager '{manager_id}' in department '{department}'.",
        )

    rows = _build_report_rows(reports)

    wb = Workbook()
    ws = wb.active
    assert ws is not None  # a freshly created Workbook always has a default active sheet
    ws.title = "Team Dashboard"

    headers = [
        "Employee", "User ID", "Status", "Current Course", "Courses Completed",
        "Completion %", "Readiness Score", "RAG", "At Risk",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="005787")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for r in rows:
        ws.append([
            _formula_safe(r["display_name"]),
            _formula_safe(r["user_id"]),
            _formula_safe(r["status"]),
            _formula_safe(r["current_course_id"]),
            r["courses_completed"],
            r["completion_rate_pct"] / 100,
            r["readiness_score"],
            r["rag_color"].upper(),
            "Yes" if r["is_at_risk"] else "No",
        ])
        row_idx = ws.max_row
        pct_cell = ws.cell(row=row_idx, column=6)
        pct_cell.number_format = "0.0%"
        score_cell = ws.cell(row=row_idx, column=7)
        score_cell.number_format = "0.00"
        rag_cell = ws.cell(row=row_idx, column=8)
        rag_cell.fill = _RAG_FILLS[r["rag_color"]]
        rag_cell.font = _RAG_FONTS[r["rag_color"]]
        rag_cell.alignment = Alignment(horizontal="center")
        if r["is_at_risk"]:
            ws.cell(row=row_idx, column=9).font = Font(color="B91C1C", bold=True)

    widths = [22, 12, 16, 18, 16, 13, 15, 10, 10]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"team-dashboard-{manager_id}-{department}.xlsx"
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Bucket 3: Strategic (Team vs Department Baseline) ────────────────────────

@router.get("/{manager_id}/strategic")
async def manager_strategic(manager_id: str, department: str = DEFAULT_DEPARTMENT, role: str = ""):
    """Strategic Bucket — compares the manager's team metrics against the department KPI baseline.

    Reads from the central KPI store so this data is PII-free and pre-aggregated.
    Falls back to department-level live calculation if no KPI payload exists yet.
    """
    _require_manager(role)
    store = DepartmentScopedStore(department)
    reports = _get_direct_reports(store, manager_id)

    if not reports:
        raise HTTPException(
            status_code=404,
            detail=f"No direct reports found for manager '{manager_id}' in department '{department}'.",
        )

    # Team metrics (live)
    total = len(reports)
    team_readiness_scores = [p.get("readiness_score", 0.0) for p in reports]
    team_avg_readiness = sum(team_readiness_scores) / total if total else 0.0
    max_courses = get_param("MAX_COURSES")
    team_completion = sum(
        len(p.get("completed_courses", [])) / max_courses for p in reports
    ) / total * 100 if total else 0.0

    # Department baseline from KPI store (today's payload, generated lazily if missing)
    payload = ensure_kpi_payload_for_today(department)
    dept_avg_readiness = None
    dept_avg_completion = None
    if payload:
        dept_avg_readiness = payload.get("risk_indicators", {}).get("avg_readiness_score")
        dept_avg_completion = payload.get("learning_metrics", {}).get("avg_completion_rate_pct")

    # If no KPI payload exists yet, fall back to all-department live calculation
    if dept_avg_readiness is None:
        all_progress = store.read_all_user_progress()
        all_scores = [p.get("readiness_score", 0.0) for p in all_progress]
        dept_avg_readiness = sum(all_scores) / len(all_scores) if all_scores else 0.0
        dept_avg_completion = (
            sum(len(p.get("completed_courses", [])) / max_courses for p in all_progress)
            / len(all_progress) * 100
            if all_progress else 0.0
        )

    readiness_delta = round(team_avg_readiness - dept_avg_readiness, 2)
    completion_delta = round(team_completion - (dept_avg_completion or 0.0), 1)

    return {
        "manager_id": manager_id,
        "department": department,
        "team": {
            "avg_readiness_score": round(team_avg_readiness, 2),
            "avg_completion_rate_pct": round(team_completion, 1),
        },
        "department_baseline": {
            "avg_readiness_score": round(dept_avg_readiness, 2),
            "avg_completion_rate_pct": round(dept_avg_completion or 0.0, 1),
        },
        "delta": {
            "readiness": readiness_delta,
            "readiness_label": "above" if readiness_delta >= 0 else "below",
            "completion_pct": completion_delta,
            "completion_label": "above" if completion_delta >= 0 else "below",
        },
    }
